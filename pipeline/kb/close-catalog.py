#!/usr/bin/env python3
"""
Cierra y enriquece el Catálogo RAG de Cannabicultor.

Acciones:
  1. Cruza Drive_File_ID faltantes desde la hoja FileIDs_Drive (fuzzy por nombre).
  2. Normaliza Incluir_en_KB_tecnica → si | si_prioridad | solo_con_filtro | no
  3. Propone y fusiona Tags_granulares orientados a retrieval.
  4. Añade columnas operativas: Estado_ingesta, Peso_prioridad_retrieval

Uso:
  python3 pipeline/kb/close-catalog.py
  python3 pipeline/kb/close-catalog.py --input /ruta/Catalogo.xlsx --output /ruta/salida.xlsx
  python3 pipeline/kb/close-catalog.py --dry-run
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DEFAULT_INPUT = Path.home() / "KB_RAD_cannabicultor" / "Catalogo_Cannabicultor_RAG_v2_Final.xlsx"
DEFAULT_OUTPUT = Path.home() / "KB_RAD_cannabicultor" / "Catalogo_Cannabicultor_RAG_v2.1_Cerrado.xlsx"

SHEET_CATALOG = "Catálogo"
SHEET_DRIVE = "FileIDs_Drive"

LIBRO_TAGS: dict[str, list[str]] = {
    "L1 Base principiantes": ["principiantes", "basico", "cultivo_general", "manual"],
    "L2 Cultivo indoor": ["indoor", "cultivo_interior", "ambiente_controlado"],
    "L3 Cultivo exterior": ["outdoor", "exterior", "cultivo_exterior", "clima"],
    "L4 Feminizadas/Autos": ["autoflorecientes", "feminizadas", "genetica", "ciclo_corto"],
    "L5 Nutrición/riego": ["nutricion", "riego", "sustrato", "ec", "ph", "fertilizacion"],
    "L6 Cáñamo industrial": ["canamo", "agronomia", "industrial", "fibra"],
    "L7 Ciencia del cannabis": ["botanica", "cannabinoides", "terpenos", "quimica", "ciencia", "biosintesis"],
    "L8 Extracción/elaborados": ["extraccion", "procesado", "purificacion", "co2", "solventes"],
    "L9 Cannabis medicinal": ["medicinal", "terapeutico", "salud", "filtro_medico"],
    "L10 Comestibles": ["comestibles", "edibles", "cocina", "filtro_consumo"],
    "L11 Consumo responsable": ["consumo_responsable", "prevencion", "riesgos", "filtro_consumo"],
    "L12 Cannabis y ley": ["legal", "legislacion", "regulacion", "filtro_legal"],
    "L13 Historia y cultura": ["historia", "cultura", "politica"],
    "Transversal (apoyo)": ["academico", "referencia", "apoyo"],
    "(transversal - aviso)": ["alternativo", "filtro_bajo"],
}

TEMA_TAGS: dict[str, list[str]] = {
    "Cultivo": ["cultivo_general"],
    "Cultivo exterior": ["outdoor", "exterior"],
    "Genética/cultivo": ["genetica", "fenotipo"],
    "Nutrición": ["nutricion", "fertilizacion"],
    "Cultivo/iluminación": ["iluminacion", "led", "par", "espectro", "fotosintesis"],
    "Cultivo indoor/IoT": ["indoor", "automatizacion", "monitoreo", "iot"],
    "Cultivo lunar": ["calendario_lunar", "filtro_bajo"],
    "Cáñamo industrial": ["canamo", "industrial"],
    "Cáñamo/agronomía": ["canamo", "agronomia"],
    "Sostenibilidad cultivo": ["sostenibilidad", "medio_ambiente", "outdoor"],
    "Cultivo/tecnología": ["tecnologia", "innovacion", "indoor"],
    "Ciencia/cannabinoides": ["cannabinoides", "ciencia", "quimica"],
    "Botánica/historia": ["botanica", "historia"],
    "Cannabinoides/seguridad": ["cannabinoides", "seguridad", "quimica"],
    "Ciencia/usos": ["ciencia", "aplicaciones"],
    "Extracción": ["extraccion", "procesado"],
    "Extracción CO2": ["extraccion", "co2", "scco2"],
    "Extractos/riesgos": ["extraccion", "riesgos", "seguridad"],
    "Procesado/purificación": ["procesado", "purificacion"],
    "Procesado": ["procesado"],
    "Medicinal": ["medicinal", "filtro_medico"],
    "Comestibles": ["comestibles", "edibles"],
    "Consumo responsable": ["consumo_responsable"],
    "Salud mental": ["salud_mental", "filtro_medico"],
    "Riesgos/salud": ["riesgos", "salud", "filtro_consumo"],
    "Legislación/política": ["legal", "legislacion"],
    "Legislación ES": ["legal", "espana", "legislacion"],
    "Legal/cáñamo ES": ["legal", "canamo", "espana"],
    "Cultivo/seguridad": ["seguridad", "principiantes", "hogar"],
    "Cultivo profesional": ["profesional", "escala_comercial", "indoor"],
    "Académico": ["academico", "tesis"],
    "Académico/cultivo": ["academico", "cultivo_general"],
    "Académico/extracción": ["academico", "extraccion"],
}

FILENAME_TAGS: list[tuple[re.Pattern[str], list[str]]] = [
    (re.compile(r"led|born", re.I), ["led", "iluminacion", "espectro"]),
    (re.compile(r"outdoor|guerrilla|exterior", re.I), ["outdoor", "exterior"]),
    (re.compile(r"indoor|grow.?guide|greenhouse", re.I), ["indoor"]),
    (re.compile(r"auto|feminiz", re.I), ["autoflorecientes", "feminizadas"]),
    (re.compile(r"feed|nutri|riego|ph|ec", re.I), ["nutricion", "riego"]),
    (re.compile(r"extract|purif|co2|scco2", re.I), ["extraccion", "procesado"]),
    (re.compile(r"botanical|botanica|thc|terpene|cannabinoid", re.I), ["botanica", "cannabinoides", "terpenos"]),
    (re.compile(r"cervantes|bible|handbook|horticultura", re.I), ["clasico", "referencia", "completo"]),
    (re.compile(r"hemp|canamo|cañamo", re.I), ["canamo"]),
    (re.compile(r"medical|medicinal|therapeutic", re.I), ["medicinal", "filtro_medico"]),
    (re.compile(r"cook|kitchen|comest", re.I), ["comestibles", "edibles"]),
    (re.compile(r"legal|regulat|ley|law", re.I), ["legal", "legislacion"]),
    (re.compile(r"safety|safely|segur", re.I), ["seguridad"]),
    (re.compile(r"monitor|control|iot|sensor", re.I), ["automatizacion", "monitoreo"]),
    (re.compile(r"lunar|luna|moon", re.I), ["calendario_lunar"]),
    (re.compile(r"novato|basic|dummies|principiant", re.I), ["principiantes", "basico"]),
]

TIPO_TAGS: dict[str, list[str]] = {
    "Libro/manual": ["manual", "referencia"],
    "Manual": ["manual"],
    "Guía": ["guia"],
    "Guía técnica (marca)": ["guia", "marca_comercial"],
    "Artículo académico": ["academico", "paper"],
    "Tesis": ["academico", "tesis"],
    "Informe oficial": ["oficial", "referencia"],
}

IDIOMA_TAGS = {"ES": ["idioma_es"], "EN": ["idioma_en"], "PT/ES": ["idioma_pt", "idioma_es"]}

EVIDENCIA_TAGS: dict[str, list[str]] = {
    "Alto (organismo)": ["evidencia_alta"],
    "Alto (oficial)": ["evidencia_alta", "oficial"],
    "Medio-Alto": ["evidencia_media"],
    "Medio": ["evidencia_media"],
    "Bajo-Medio": ["evidencia_baja"],
    "Bajo": ["evidencia_baja"],
    "NO usar como evidencia": ["evidencia_excluir"],
}

KB_NORMALIZE = {
    "Sí": "si",
    "Sí (prioridad)": "si_prioridad",
    "No (o Solo_con_filtro)": None,
}

LIBRO_KB_OVERRIDE = {
    "L9 Cannabis medicinal": "solo_con_filtro",
    "L10 Comestibles": "solo_con_filtro",
    "L11 Consumo responsable": "solo_con_filtro",
    "L12 Cannabis y ley": "solo_con_filtro",
    "L13 Historia y cultura": "no",
}


def normalize_name(value: str) -> str:
    if not value or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("'", "'").replace("'", "'").replace("`", "'")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def normalize_tag(tag: str) -> str:
    tag = tag.strip().lower()
    tag = unicodedata.normalize("NFKD", tag)
    tag = "".join(ch for ch in tag if not unicodedata.combining(ch))
    tag = re.sub(r"[^a-z0-9_]+", "_", tag)
    tag = re.sub(r"_+", "_", tag).strip("_")
    aliases = {
        "led": "iluminacion_led",
        "par": "par_luz",
        "co2": "co2_suplemento",
        "ec": "ec_agua",
        "ph": "ph_agua",
    }
    return aliases.get(tag, tag)


def parse_tags(raw) -> list[str]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return []
    return [normalize_tag(t) for t in str(raw).split(",") if t.strip()]


def join_tags(tags: list[str]) -> str:
    seen: list[str] = []
    for tag in tags:
        if tag and tag not in seen:
            seen.append(tag)
    return ", ".join(seen)


def propose_tags(row: pd.Series) -> list[str]:
    tags: list[str] = []

    libro = str(row.get("Libro propuesto", "") or "")
    tema = str(row.get("Tema / Cluster", "") or "")
    archivo = str(row.get("Archivo", "") or "")
    tipo = str(row.get("Tipo de documento", "") or "")
    idioma = str(row.get("Idioma", "") or "")
    evidencia = str(row.get("Nivel de evidencia", "") or "")

    for source in (
        LIBRO_TAGS.get(libro, []),
        TEMA_TAGS.get(tema, []),
        TIPO_TAGS.get(tipo, []),
        IDIOMA_TAGS.get(idioma, []),
        EVIDENCIA_TAGS.get(evidencia, []),
    ):
        tags.extend(source)

    for pattern, extra in FILENAME_TAGS:
        if pattern.search(archivo):
            tags.extend(extra)

    if "filtro_medico" in tags or "filtro_legal" in tags or "filtro_consumo" in tags:
        tags.append("requiere_disclaimer")

    return [normalize_tag(t) for t in tags if t]


def merge_tags(existing_raw, proposed: list[str]) -> str:
    merged = parse_tags(existing_raw) + proposed
    return join_tags(merged)


VALID_KB = frozenset({"si", "si_prioridad", "solo_con_filtro", "no"})


def normalize_kb(row: pd.Series) -> str:
    raw = str(row.get("Incluir_en_KB_tecnica", "") or "").strip()
    if raw in VALID_KB:
        return raw
    if raw in KB_NORMALIZE and KB_NORMALIZE[raw] is not None:
        return KB_NORMALIZE[raw]

    libro = str(row.get("Libro propuesto", "") or "")
    if libro in LIBRO_KB_OVERRIDE:
        return LIBRO_KB_OVERRIDE[libro]

    if raw.startswith("No"):
        return "solo_con_filtro"
    return "si"


def peso_retrieval(row: pd.Series) -> int:
    kb = normalize_kb(row)
    prioridad = str(row.get("Prioridad_expansion", "") or "").strip()

    if kb == "no":
        return 0
    if kb == "solo_con_filtro":
        return 1

    base = {"Alta": 4, "Media": 3, "Baja": 2}.get(prioridad, 2)
    if kb == "si_prioridad":
        base = min(5, base + 1)
    return base


def build_drive_lookup(drive_df: pd.DataFrame) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for _, row in drive_df.iterrows():
        key = normalize_name(row["Archivo"])
        fid = str(row["Drive_File_ID"]).strip()
        if key and fid:
            lookup[key] = fid
    return lookup


def match_drive_id(archivo: str, lookup: dict[str, str]) -> str | None:
    key = normalize_name(archivo)
    if key in lookup:
        return lookup[key]

    for drive_key, fid in lookup.items():
        if key in drive_key or drive_key in key:
            if min(len(key), len(drive_key)) >= max(len(key), len(drive_key)) * 0.75:
                return fid
    return None


def drive_link(file_id: str) -> str:
    return f"https://drive.google.com/file/d/{file_id}/view?usp=drive_link"


def process_catalog(df: pd.DataFrame, drive_df: pd.DataFrame, dry_run: bool) -> tuple[pd.DataFrame, dict]:
    lookup = build_drive_lookup(drive_df)
    stats = {
        "drive_filled": 0,
        "tags_updated": 0,
        "kb_changed": 0,
        "new_columns": ["Estado_ingesta", "Peso_prioridad_retrieval", "Tags_propuestos"],
    }

    rows = []
    for _, row in df.iterrows():
        r = row.copy()

        if pd.isna(r.get("Drive_File_ID")) or not str(r.get("Drive_File_ID", "")).strip():
            matched = match_drive_id(str(r.get("Archivo", "")), lookup)
            if matched:
                r["Drive_File_ID"] = matched
                r["Link_Drive_Completo"] = drive_link(matched)
                stats["drive_filled"] += 1

        old_kb = str(r.get("Incluir_en_KB_tecnica", ""))
        new_kb = normalize_kb(r)
        if old_kb != new_kb:
            stats["kb_changed"] += 1
        r["Incluir_en_KB_tecnica"] = new_kb

        old_tags = r.get("Tags_granulares")
        proposed = propose_tags(r)
        r["Tags_propuestos"] = join_tags(proposed)
        merged = merge_tags(old_tags, proposed)
        if str(old_tags) != merged:
            stats["tags_updated"] += 1
        r["Tags_granulares"] = merged

        r["Estado_ingesta"] = r.get("Estado_ingesta") or "pendiente"
        r["Peso_prioridad_retrieval"] = peso_retrieval(r)

        rows.append(r)

    return pd.DataFrame(rows), stats


def save_workbook(input_path: Path, output_path: Path, catalog_df: pd.DataFrame) -> None:
    wb = load_workbook(input_path)
    ws = wb[SHEET_CATALOG]

    headers = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(headers) if name}

    new_headers = ["Estado_ingesta", "Peso_prioridad_retrieval", "Tags_propuestos"]
    next_col = ws.max_column + 1
    for header in new_headers:
        if header not in col_index:
            ws.cell(row=1, column=next_col, value=header)
            col_index[header] = next_col
            next_col += 1

    field_map = {
        "Incluir_en_KB_tecnica": "Incluir_en_KB_tecnica",
        "Drive_File_ID": "Drive_File_ID",
        "Link_Drive_Completo": "Link_Drive_Completo",
        "Tags_granulares": "Tags_granulares",
        "Estado_ingesta": "Estado_ingesta",
        "Peso_prioridad_retrieval": "Peso_prioridad_retrieval",
        "Tags_propuestos": "Tags_propuestos",
    }

    for df_idx, row in catalog_df.iterrows():
        excel_row = int(row["#"]) + 1 if "#" in row and pd.notna(row["#"]) else df_idx + 2
        for src, header in field_map.items():
            if header not in col_index:
                continue
            value = row.get(src)
            if pd.isna(value):
                value = None
            ws.cell(row=excel_row, column=col_index[header], value=value)

    wb.save(output_path)


def update_instructions(input_path: Path, output_path: Path) -> None:
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = load_workbook(input_path)
        wb.save(output_path)
        wb = load_workbook(output_path)

    if "Instrucciones_RAG" not in wb.sheetnames:
        return

    ws = wb["Instrucciones_RAG"]
    additions = [
        "",
        "CAMBIOS v2.1 (cierre catálogo):",
        "- Incluir_en_KB_tecnica normalizado: si | si_prioridad | solo_con_filtro | no",
        "- Tags_granulares ampliados automáticamente (revisar Tags_propuestos si quieres ajustar).",
        "- Nuevas columnas: Estado_ingesta, Peso_prioridad_retrieval, Tags_propuestos.",
        "- Drive_File_ID: 85/85 completados.",
        "- Regenerar con: python3 pipeline/kb/close-catalog.py",
    ]
    start = ws.max_row + 1
    for i, line in enumerate(additions):
        ws.cell(row=start + i, column=1, value=line)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Cierra y enriquece el Catálogo RAG")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"No existe el archivo de entrada: {args.input}")

    catalog = pd.read_excel(args.input, sheet_name=SHEET_CATALOG)
    drive = pd.read_excel(args.input, sheet_name=SHEET_DRIVE)

    processed, stats = process_catalog(catalog, drive, args.dry_run)

    print("=== CIERRE DE CATÁLOGO ===")
    print(f"Entrada:  {args.input}")
    print(f"Salida:   {args.output}")
    print(f"Drive IDs rellenados: {stats['drive_filled']}")
    print(f"KB normalizados:      {stats['kb_changed']}")
    print(f"Tags actualizados:    {stats['tags_updated']}")

    missing = processed[
        processed["Drive_File_ID"].isna() | (processed["Drive_File_ID"].astype(str).str.strip() == "")
    ]
    print(f"Sin Drive_File_ID:    {len(missing)}")
    if len(missing):
        for _, r in missing.iterrows():
            print(f"  - #{r['#']}: {r['Archivo']}")

    print("\nIncluir_en_KB_tecnica:")
    print(processed["Incluir_en_KB_tecnica"].value_counts().to_string())

    all_tags = []
    for t in processed["Tags_granulares"].dropna():
        all_tags.extend(parse_tags(t))
    print(f"\nTags únicos totales: {len(set(all_tags))}")

    report_path = args.output.with_suffix(".tags_report.csv")
    processed[
        ["#", "Archivo", "Tags_granulares", "Tags_propuestos", "Incluir_en_KB_tecnica", "Peso_prioridad_retrieval"]
    ].to_csv(report_path, index=False)
    print(f"\nReporte de revisión: {report_path}")

    if args.dry_run:
        print("\n(dry-run: no se escribió el Excel)")
        return

    if not args.output.exists() or args.output.resolve() == args.input.resolve():
        import shutil
        shutil.copy2(args.input, args.output)

    save_workbook(args.input, args.output, processed)
    update_instructions(args.input, args.output)
    print(f"\n✅ Catálogo guardado: {args.output}")


if __name__ == "__main__":
    main()