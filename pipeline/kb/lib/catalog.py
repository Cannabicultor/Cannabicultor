from __future__ import annotations

from dataclasses import dataclass, field, asdict
from pathlib import Path

import pandas as pd

SHEET_CATALOG = "Catálogo"
KB_INCLUDE = {"si", "si_prioridad"}


@dataclass
class CatalogDocument:
    catalog_num: int
    archivo: str
    drive_file_id: str
    link_drive: str | None
    idioma_contenido: str
    politica_idioma: str | None
    factor_idioma_retrieval: float
    incluir_en_kb: str
    peso_prioridad_retrieval: int
    libro_propuesto: str | None
    tema_cluster: str | None
    tipo_documento: str | None
    nivel_evidencia: str | None
    prioridad_expansion: str | None
    tags: list[str] = field(default_factory=list)
    notas_rag: str | None = None
    estado_ingesta: str = "pendiente"

    @property
    def respuesta_requiere_traduccion(self) -> bool:
        return (
            self.idioma_contenido == "en"
            or "respuesta_requiere_traduccion" in self.tags
        )

    def snapshot(self) -> dict:
        return asdict(self)


def parse_tags(raw) -> list[str]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return []
    return [t.strip().lower() for t in str(raw).split(",") if t.strip()]


def load_catalog(path: Path, only_kb: bool = True) -> list[CatalogDocument]:
    df = pd.read_excel(path, sheet_name=SHEET_CATALOG)
    docs: list[CatalogDocument] = []

    for _, row in df.iterrows():
        incluir = str(row.get("Incluir_en_KB_tecnica", "") or "").strip()
        if only_kb and incluir not in KB_INCLUDE:
            continue

        drive_id = str(row.get("Drive_File_ID", "") or "").strip()
        if not drive_id:
            raise ValueError(f"Doc #{row['#']} sin Drive_File_ID: {row.get('Archivo')}")

        peso = row.get("Peso_prioridad_retrieval")
        if pd.isna(peso):
            peso = 2
        factor = row.get("factor_idioma_retrieval")
        if pd.isna(factor):
            factor = 1.0

        docs.append(
            CatalogDocument(
                catalog_num=int(row["#"]),
                archivo=str(row["Archivo"]).strip(),
                drive_file_id=drive_id,
                link_drive=str(row.get("Link_Drive_Completo") or "") or None,
                idioma_contenido=str(row.get("idioma_contenido") or "es").strip().lower(),
                politica_idioma=str(row.get("politica_idioma") or "") or None,
                factor_idioma_retrieval=float(factor),
                incluir_en_kb=incluir,
                peso_prioridad_retrieval=int(peso),
                libro_propuesto=str(row.get("Libro propuesto") or "") or None,
                tema_cluster=str(row.get("Tema / Cluster") or "") or None,
                tipo_documento=str(row.get("Tipo de documento") or "") or None,
                nivel_evidencia=str(row.get("Nivel de evidencia") or "") or None,
                prioridad_expansion=str(row.get("Prioridad_expansion") or "") or None,
                tags=parse_tags(row.get("Tags_granulares")),
                notas_rag=str(row.get("Notas_RAG") or "") or None,
                estado_ingesta=str(row.get("Estado_ingesta") or "pendiente"),
            )
        )

    return sorted(docs, key=lambda d: d.catalog_num)


def update_catalog_estado(path: Path, statuses: dict[int, str]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb[SHEET_CATALOG]
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
    col = headers.get("Estado_ingesta")
    col_num = headers.get("#")
    if not col or not col_num:
        wb.save(path)
        return

    for row_idx in range(2, ws.max_row + 1):
        num = ws.cell(row=row_idx, column=col_num).value
        if num is None:
            continue
        num = int(num)
        if num in statuses:
            ws.cell(row=row_idx, column=col, value=statuses[num])

    wb.save(path)