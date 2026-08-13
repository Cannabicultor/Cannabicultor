#!/usr/bin/env python3
"""
Aplica la estrategia multilingüe definitiva al Catálogo RAG → v2.2

Columnas nuevas (mínimas):
  - idioma_contenido
  - politica_idioma
  - factor_idioma_retrieval

Uso:
  python3 pipeline/kb/apply-multilingual.py
"""

from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DEFAULT_INPUT = Path.home() / "KB_RAD_cannabicultor" / "Catalogo_Cannabicultor_RAG_v2.1_Cerrado.xlsx"
DEFAULT_OUTPUT = Path.home() / "KB_RAD_cannabicultor" / "Catalogo_Cannabicultor_RAG_v2.2_Multilingual.xlsx"

SHEET_CATALOG = "Catálogo"

# Documentos EN redundantes o de bajo valor → fuera del corpus técnico
EXCLUDE_FROM_KB: dict[int, str] = {
    3: "L1 EN redundante (hay 6 manuales ES equivalentes)",
    4: "L1 EN redundante (Grower's Handbook duplica contenido ES)",
    16: "Guía genérica convertida, baja calidad editorial",
    18: "Duplicado de Grow-Guide.pdf (#19)",
    19: "Duplicado genérico indoor (ya hay Berger #15 y guías ES)",
    82: "Safety guide EN genérico; no aporta a cultivador hispano",
    83: "Duplicado safety EN; contexto anglosajón",
}

# Política explícita para documentos EN que permanecen en KB
EN_POLITICA: dict[int, str] = {
    12: "es_prioritario",
    13: "en_aceptado",
    14: "es_prioritario",
    15: "en_aceptado",
    17: "es_prioritario",
    20: "en_aceptado",
    22: "en_aceptado",
    24: "en_prioritario",
    34: "en_prioritario",
    37: "en_prioritario",
    38: "en_prioritario",
    40: "en_prioritario",
    41: "en_prioritario",
    42: "en_prioritario",
}

FACTOR_BY_POLITICA_ES = 1.0
FACTOR_BY_POLITICA: dict[str, float] = {
    "es_obligatorio": 1.0,
    "es_prioritario": 1.0,
    "en_aceptado": 0.80,
    "en_prioritario": 0.85,
    "excluir_en": 0.0,
}

ES_OBLIGATORIO_LIBROS = {
    "L1 Base principiantes",
    "L6 Cáñamo industrial",
}

TAG_ALIASES = {
    "idioma_es": "lang_es",
    "idioma_en": "lang_en",
    "idioma_pt": "lang_pt",
}

STRATEGY_LINES = [
    "",
    "ESTRATEGIA MULTILINGÜE v2.2 (definitiva para ingesta):",
    "",
    "PRINCIPIO: Salida siempre en español. Inglés = evidencia técnica interna.",
    "",
    "RETRIEVAL:",
    "  score = similitud × Peso_prioridad_retrieval × factor_idioma_retrieval",
    "  - Embeddings multilingües (ES/EN en el mismo espacio vectorial).",
    "  - Sin filtro duro de idioma salvo queries con tag principiantes → solo lang_es.",
    "  - factor_idioma_retrieval: ES=1.0 | en_prioritario=0.85 | en_aceptado=0.80 | es_prioritario+EN=0.75",
    "",
    "GENERACIÓN:",
    "  - Chunks EN nunca se muestran al usuario en inglés.",
    "  - El LLM sintetiza en español con voz Cannabicultor.",
    "  - Tag respuesta_requiere_traduccion en docs EN del corpus.",
    "",
    "POLÍTICA POR CLUSTER:",
    "  L7/L5/L8: EN activo (en_prioritario / en_aceptado)",
    "  L1/L6: solo ES (es_obligatorio)",
    "  L2/L3/L4: ES primero; EN selectivo",
    "",
    "CORPUS TÉCNICO: ver hoja Catálogo, filtro Incluir_en_KB_tecnica ∈ {si, si_prioridad}",
    "Regenerar: python3 pipeline/kb/apply-multilingual.py",
]


def normalize_idioma_contenido(raw: str) -> str:
    val = str(raw or "").strip().upper()
    if val == "EN":
        return "en"
    if val in {"PT/ES", "PT-ES", "MIXTO"}:
        return "mixto"
    return "es"


def parse_tags(raw) -> list[str]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return []
    return [t.strip() for t in str(raw).split(",") if t.strip()]


def join_tags(tags: list[str]) -> str:
    seen: list[str] = []
    for tag in tags:
        if tag and tag not in seen:
            seen.append(tag)
    return ", ".join(seen)


def standardize_lang_tags(tags: list[str], idioma: str, politica: str, in_kb: bool) -> list[str]:
    cleaned: list[str] = []
    for tag in tags:
        low = tag.strip().lower()
        if low in TAG_ALIASES:
            low = TAG_ALIASES[low]
        if low in {"lang_es", "lang_en", "lang_pt", "lang_mixto"}:
            continue
        cleaned.append(low)

    lang_tag = {
        "es": "lang_es",
        "en": "lang_en",
        "mixto": "lang_mixto",
    }.get(idioma, "lang_es")
    cleaned.append(lang_tag)
    if idioma == "mixto":
        cleaned.append("lang_pt")

    if in_kb and idioma == "en":
        cleaned.append("respuesta_requiere_traduccion")
        if politica == "en_prioritario":
            cleaned.append("evidencia_internacional")

    if in_kb and idioma == "es" and politica == "es_obligatorio":
        cleaned.append("contexto_hispano")

    return cleaned


def infer_politica(row: pd.Series, doc_num: int, in_kb: bool) -> str:
    if doc_num in EXCLUDE_FROM_KB:
        return "excluir_en"

    idioma = row["idioma_contenido"]
    libro = str(row.get("Libro propuesto", "") or "")

    if doc_num in EN_POLITICA:
        return EN_POLITICA[doc_num]

    if not in_kb:
        if idioma == "en":
            return "en_aceptado"
        return "es_prioritario"

    if idioma == "es" and libro in ES_OBLIGATORIO_LIBROS:
        return "es_obligatorio"

    if idioma == "es":
        return "es_prioritario"

    if idioma == "en":
        return "en_aceptado"

    return "es_prioritario"


def infer_factor(idioma: str, politica: str, in_kb: bool) -> float:
    if not in_kb or politica == "excluir_en":
        return 0.0
    if idioma == "es":
        return FACTOR_BY_POLITICA_ES
    if politica == "es_prioritario":
        return 0.75
    return FACTOR_BY_POLITICA.get(politica, 0.80)


def in_technical_kb(kb_val: str) -> bool:
    return str(kb_val).strip() in {"si", "si_prioridad"}


def process(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    stats = {
        "excluded": [],
        "en_kept": [],
        "corpus_before": int(df["Incluir_en_KB_tecnica"].isin(["si", "si_prioridad"]).sum()),
    }

    rows = []
    for _, row in df.iterrows():
        r = row.copy()
        doc_num = int(r["#"])
        idioma = normalize_idioma_contenido(r.get("Idioma"))
        r["idioma_contenido"] = idioma

        kb = str(r.get("Incluir_en_KB_tecnica", "") or "").strip()
        if doc_num in EXCLUDE_FROM_KB:
            old_kb = kb
            r["Incluir_en_KB_tecnica"] = "no"
            kb = "no"
            nota = str(r.get("Notas_RAG") or "")
            reason = EXCLUDE_FROM_KB[doc_num]
            extra = f"[v2.2] Excluido corpus: {reason}"
            r["Notas_RAG"] = f"{nota} | {extra}".strip(" |") if nota else extra
            stats["excluded"].append({"#": doc_num, "archivo": r["Archivo"], "antes": old_kb, "motivo": reason})

        in_kb = in_technical_kb(kb)
        politica = infer_politica(r, doc_num, in_kb)
        r["politica_idioma"] = politica
        r["factor_idioma_retrieval"] = infer_factor(idioma, politica, in_kb)

        tags = standardize_lang_tags(parse_tags(r.get("Tags_granulares")), idioma, politica, in_kb)
        r["Tags_granulares"] = join_tags(tags)

        if in_kb and idioma == "en":
            stats["en_kept"].append({"#": doc_num, "archivo": r["Archivo"], "politica": politica})

        rows.append(r)

    out = pd.DataFrame(rows)
    stats["corpus_after"] = int(out["Incluir_en_KB_tecnica"].isin(["si", "si_prioridad"]).sum())
    stats["en_in_corpus"] = int(
        out[out["Incluir_en_KB_tecnica"].isin(["si", "si_prioridad"]) & (out["idioma_contenido"] == "en")].shape[0]
    )
    return out, stats


def save_workbook(template: Path, output: Path, df: pd.DataFrame) -> None:
    shutil.copy2(template, output)
    wb = load_workbook(output)
    ws = wb[SHEET_CATALOG]

    headers = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(headers) if name}

    new_cols = ["idioma_contenido", "politica_idioma", "factor_idioma_retrieval"]
    next_col = ws.max_column + 1
    for header in new_cols:
        if header not in col_index:
            ws.cell(row=1, column=next_col, value=header)
            col_index[header] = next_col
            next_col += 1

    fields = [
        "Incluir_en_KB_tecnica",
        "Tags_granulares",
        "Notas_RAG",
        "idioma_contenido",
        "politica_idioma",
        "factor_idioma_retrieval",
    ]

    for _, row in df.iterrows():
        excel_row = int(row["#"]) + 1
        for field in fields:
            if field not in col_index:
                continue
            value = row.get(field)
            if pd.isna(value):
                value = None
            ws.cell(row=excel_row, column=col_index[field], value=value)

    if "Instrucciones_RAG" in wb.sheetnames:
        ws_i = wb["Instrucciones_RAG"]
        start = ws_i.max_row + 1
        for i, line in enumerate(STRATEGY_LINES):
            ws_i.cell(row=start + i, column=1, value=line)

    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"No existe: {args.input}")

    df = pd.read_excel(args.input, sheet_name=SHEET_CATALOG)
    processed, stats = process(df)

    save_workbook(args.input, args.output, processed)

    report = processed[
        processed["Incluir_en_KB_tecnica"].isin(["si", "si_prioridad"])
    ][
        [
            "#",
            "Archivo",
            "idioma_contenido",
            "politica_idioma",
            "factor_idioma_retrieval",
            "Incluir_en_KB_tecnica",
            "Peso_prioridad_retrieval",
            "Libro propuesto",
        ]
    ].sort_values("#")

    report_path = args.output.with_suffix(".corpus_v2.2.csv")
    report.to_csv(report_path, index=False)

    excluded_path = args.output.with_suffix(".excluded_v2.2.csv")
    pd.DataFrame(stats["excluded"]).to_csv(excluded_path, index=False)

    print("=== v2.2 MULTILINGUAL ===")
    print(f"Salida: {args.output}")
    print(f"Corpus: {stats['corpus_before']} → {stats['corpus_after']} docs")
    print(f"EN en corpus: {stats['en_in_corpus']}")
    print(f"Excluidos: {len(stats['excluded'])}")
    print(f"Reporte corpus: {report_path}")
    print(f"Reporte excluidos: {excluded_path}")


if __name__ == "__main__":
    main()